#!/usr/bin/python3
# -*- coding: utf-8 -*-
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from pages.OMPage import OMPage


url = 'http://om.motionglobal.com'
driver = webdriver.Firefox()
driver.implicitly_wait(10)
driver.maximize_window()
driver.set_window_size(1440, 900)
driver.get(url)
action = ActionChains(driver)


def login():
    om_page = OMPage(driver)
    om_page.login()


def transfer_order(order_code):
    driver.get('http://om.motionglobal.com/salesorder/index/')
    WebDriverWait(driver, 30, 0.5).until(EC.presence_of_element_located((By.ID, 'order_id')))
    order_input = driver.find_element_by_id('order_id')
    order_input.clear()
    order_input.send_keys(order_code)
    go = driver.find_element_by_id('button')
    go.click()
    ok = driver.find_element_by_xpath('//div[@id="remoteBox-footer"]/input[1]')
    try:
        if EC.visibility_of(ok):
            ok.click()
            WebDriverWait(driver, 30, 0.5).until(EC.presence_of_element_located(
                (By.XPATH, '/html/body/div[2]/div[3]/div/div/div[4]/div[2]/form/table/tbody/tr/td[2]/div/a')))
        else:
            pass
    except Exception as e:
        print("order had been transfered" % e)


def cancel_order(order_code):
    driver.get("http://om.motionglobal.com/salesorder/edit/?oid=" + order_code)
    cancel_radio = driver.find_element_by_id('check-all-items')
    WebDriverWait(driver, 30, 0.5).until(EC.presence_of_element_located((By.ID, 'check-all-items')))
    cancel_radio.click()
    cancel_button = driver.find_element_by_id('button4')
    WebDriverWait(driver, 30, 0.5).until(EC.presence_of_element_located((By.ID, 'button4')))
    cancel_button.click()
    sleep(2)
    confirm_yes = driver.find_element_by_xpath("//div[@id='return_form']/table/tbody/tr[6]/td[2]/input[@id='button6']")
    WebDriverWait(driver, 30, 0.5).until(EC.presence_of_element_located(
        (By.XPATH, "//div[@id='return_form']/table/tbody/tr[6]/td[2]/input[@id='button6']")))
    confirm_yes.click()


def get_order_code(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    array = []
    for i in range(1, sheet.max_row + 1, 1):
        order_code = sheet.cell(row=i, column=1).value
        array.append(order_code)
    return array


def bath_transfer_order(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    for i in range(1, sheet.max_row + 1, 1):
        order_code = sheet.cell(row=i, column=1).value
        if order_code is None:
            break
        else:
            transfer_order(order_code)
            print(order_code + ' :' + 'transfer successfully')


def bath_cancel_order(excel_path, sheet_name):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    for i in range(1, sheet.max_row + 1, 1):
        order_code = sheet.cell(row=i, column=1).value
        if order_code is None:
            break
        else:
            cancel_order(order_code)
            print(order_code + ' :' + 'cancel successfully')


if __name__ == '__main__':
    login()
    # bath_transfer_order('/Users/admin/test/cancel_order.xlsx', 'Sheet1')
    bath_cancel_order('/Users/admin/test/cancel_order.xlsx', 'Sheet1')
    driver.quit()
