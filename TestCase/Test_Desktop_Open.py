#!/usr/bin/python
# -*- coding: <utf-8> -*-
from util.BrowserDriver import BrowserDriver
import unittest
import openpyxl
from util.BasePage import BasePage
from util.logger import Logger

logger = Logger(logger="DesktopHomepageTest").getlog()


class DesktopHomepageTest(unittest.TestCase):

    @classmethod
    def setUpClass(cls):
        driver = BrowserDriver(cls)
        cls.driver = driver.open_browser(cls)

    def setUp(self):
        pass

    def test_homepage(self):
        excel_path = "/Users/admin/test/website.xlsx"
        sheet_name = "Sheet1"
        wb = openpyxl.load_workbook(excel_path)
        sheet = wb[sheet_name]
        for i in range(1, sheet.max_row + 1, 1):
            web_url = sheet.cell(row=i, column=1).value
            if web_url is None:
                break
            else:
                self.driver.get(web_url)
                title = self.driver.title
                if "Error" in title:
                    logger.info(web_url + " " + "ERROR")
                    BasePage().get_screent_img()
                else:
                    logger.info(web_url + " " + "homepage is OK")

    def assertTitle(self, title):
        try:
            self.assertEqual(title, self.driver.title)
        except AssertionError as e:
            self.verificationErrors.append(str(e))

    @classmethod
    def tearDownClass(cls):
        cls.driver.quit()

    def tearDown(self):
        self.driver.quit()
        self.assertEqual([], self.verificationErrors)


if __name__ == "__main__":
    unittest.main()
