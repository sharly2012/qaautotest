import configparser
import yaml
import openpyxl
import csv


class BaseUtil:
    def __init__(self):
        self.root_path = '/Users/admin/PycharmProjects/SBG2018/'

    def set_root_path(self, path):
        """设置项目的主path"""
        self.root_path = path

    def get_root_path(self):
        """获取项目的主path"""
        return self.root_path

    def get_config_value(self, section, key):
        """获取config文件中section的key值"""
        config_path = self.root_path + "/conf/config.ini"
        config = configparser.ConfigParser()
        config.read(config_path)
        value = config.get(section, key)
        return value

    def get_yaml_value(self, option, key):
        """获取yaml配置文件中option的key值"""
        yaml_path = self.root_path + "/yaml/browser.yaml"
        with open(yaml_path, 'r') as f:
            temp = yaml.load(f.read())
        value = temp[option][key]
        return value

    def get_excel_data(self, excel_name, sheet_name, column_num):
        """返回excel中第column_num列的值"""
        wb = openpyxl.load_workbook(self.root_path + "/files/" + excel_name)
        sheet = wb[sheet_name]
        array = []
        # 从Excel第二行开始，第一行为列名
        for i in range(2, sheet.max_row + 1):
            date = sheet.cell(row=i, column=int(column_num)).value
            if date is None:
                break
            else:
                array.append(date)
        return array

    def get_every_row_data(self, excel_name, sheet_name):
        """获取Excel中每一行的值，返回一个二维数组"""
        wb = openpyxl.load_workbook(self.root_path + "/files/" + excel_name)
        sheet = wb[sheet_name]
        array = []
        for r in range(1, sheet.max_row + 1):
            date = (''.join([str(sheet.cell(row=r, column=c).value).ljust(10) for c in range(1, sheet.max_column + 1)]))
            if date.split()[0] == 'None':
                break
            else:
                array.append(date.split())
        return array

    def get_csv_data(self, csv_name):
        """返回一个二维数组"""
        with open(self.root_path + "/files/" + csv_name) as csv_file:
            read_csv = csv.reader(csv_file, delimiter=',')
            array = []
            for row in read_csv:
                if len(row):
                    array.append(row)
                else:
                    break
        return array
